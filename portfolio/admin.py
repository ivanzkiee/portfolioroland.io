from django.contrib import admin
from django.db.models import Count
from django.http import HttpResponse
from django.utils import timezone
from .models import ResumeDownload


@admin.action(description='Export selected resume downloads to Excel')
def export_resume_downloads(modeladmin, request, queryset):
	from openpyxl import Workbook
	from openpyxl.styles import Alignment, Font
	from openpyxl.utils import get_column_letter

	workbook = Workbook()
	sheet = workbook.active
	sheet.title = 'Resume Downloads'
	columns = ('ID', 'Full Name', 'Company', 'Position', 'Email', 'Message', 'Download Date', 'Download Time', 'Browser', 'Device', 'Operating System', 'IP Address', 'Country', 'City', 'Referrer', 'Anonymous')
	sheet.append(columns)
	for download in queryset:
		sheet.append((download.id, download.full_name, download.company, download.position, download.work_email, download.message, download.download_datetime.date(), download.download_datetime.time().replace(microsecond=0), download.browser, download.device_type, download.operating_system, download.ip_address or '', download.country, download.city, download.referrer, download.is_anonymous))

	summary = workbook.create_sheet('Analytics Summary')
	summary.append(('Metric', 'Value'))
	now = timezone.now()
	metrics = (
		('Total Downloads', queryset.count()),
		('Recruiter Downloads', queryset.filter(is_anonymous=False).count()),
		('Anonymous Downloads', queryset.filter(is_anonymous=True).count()),
		('Desktop Users', queryset.filter(device_type='Desktop').count()),
		('Mobile Users', queryset.filter(device_type='Mobile').count()),
		('Tablet Users', queryset.filter(device_type='Tablet').count()),
		('Chrome Users', queryset.filter(browser='Chrome').count()),
		('Edge Users', queryset.filter(browser='Edge').count()),
		('Firefox Users', queryset.filter(browser='Firefox').count()),
		('Downloads Today', queryset.filter(download_datetime__date=now.date()).count()),
		('Downloads This Month', queryset.filter(download_datetime__year=now.year, download_datetime__month=now.month).count()),
		('Latest Download', queryset.order_by('-download_datetime').first().download_datetime if queryset.exists() else ''),
	)
	for metric in metrics:
		summary.append(metric)
	for current_sheet in (sheet, summary):
		current_sheet.freeze_panes = 'A2'
		for cell in current_sheet[1]:
			cell.font = Font(bold=True)
			cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
		for column in current_sheet.columns:
			width = min(max(max(len(str(cell.value or '')) for cell in column) + 2, 12), 40)
			current_sheet.column_dimensions[get_column_letter(column[0].column)].width = width
		for row in current_sheet.iter_rows(min_row=2):
			for cell in row:
				cell.alignment = Alignment(vertical='top', wrap_text=True)
	for cell in sheet['G'][1:]:
		cell.number_format = 'yyyy-mm-dd'
	for cell in sheet['H'][1:]:
		cell.number_format = 'hh:mm:ss'

	response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
	response['Content-Disposition'] = f'attachment; filename="Resume_Downloads_{now:%Y-%m-%d}.xlsx"'
	workbook.save(response)
	return response


@admin.register(ResumeDownload)
class ResumeDownloadAdmin(admin.ModelAdmin):
	list_display = ('full_name', 'company', 'work_email', 'device_type', 'browser', 'download_datetime', 'is_anonymous')
	search_fields = ('full_name', 'company', 'work_email', 'position', 'browser', 'operating_system')
	list_filter = ('download_datetime', 'is_anonymous', 'browser', 'device_type', 'company')
	readonly_fields = tuple(field.name for field in ResumeDownload._meta.fields)
	actions = (export_resume_downloads,)
	change_list_template = 'admin/portfolio/resumedownload/change_list.html'

	def changelist_view(self, request, extra_context=None):
		now = timezone.localtime()
		queryset = ResumeDownload.objects.all()
		extra_context = extra_context or {}
		extra_context['resume_analytics'] = {
			'Total Downloads': queryset.count(),
			'Recruiter Downloads': queryset.filter(is_anonymous=False).count(),
			'Anonymous Downloads': queryset.filter(is_anonymous=True).count(),
			'Downloads Today': queryset.filter(download_datetime__date=now.date()).count(),
			'Downloads This Month': queryset.filter(download_datetime__year=now.year, download_datetime__month=now.month).count(),
			'Desktop Users': queryset.filter(device_type='Desktop').count(),
			'Mobile Users': queryset.filter(device_type='Mobile').count(),
			'Tablet Users': queryset.filter(device_type='Tablet').count(),
			'Top Browser': queryset.values('browser').annotate(total=Count('id')).order_by('-total').first() or {'browser': 'N/A', 'total': 0},
			'Latest Download': queryset.first(),
		}
		return super().changelist_view(request, extra_context=extra_context)

